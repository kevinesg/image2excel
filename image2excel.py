import openpyxl
import matplotlib
import pandas
import PIL
import os
import sys

from matplotlib.image import imread
from scipy.cluster.vq import whiten
from scipy.cluster.vq import kmeans


# image file input
imageFileName = sys.argv[1]
imageFile = PIL.Image.open(imageFileName)

width, height = imageFile.size # dimensions of the whole image

# input for number of rows
numRows = int(sys.argv[2])


# solve the proportional number of columns (maintain aspect ratio)
numColumns = int(numRows / height * width)


# create excel file
wb = openpyxl.Workbook()
sheet = wb.active


# set variables for dimensions of a group of pixels
group_x = width / numColumns
group_y = height / numRows # 1 group is equivalent to 1 excel cell


def std(x): # to avoid using 0 standard deviation later
    if x == 0:
        return 1
    else:
        return x


counter = 1 # counter for printing
for i in range(numRows):
    # adjust cell row height size
    sheet.row_dimensions[i + 1].height = 1080 / numRows * 0.525
    row_height = sheet.row_dimensions[i + 1].height

    for j in range(numColumns):

        print(f'Processing excel cell #{counter}...')
        counter += 1

        # convert each rectangular cell into a square
        if i == 0:
            columnLetter = openpyxl.utils.get_column_letter(j + 1)
            sheet.column_dimensions[columnLetter].width = (row_height / 6)


        # crop the pixel group to be converted to a cell
        pixelGroup = imageFile.crop((
            (j * group_x), (i * group_y),
            ((j + 1) * group_x), ((i + 1) * group_y)
            ))

        # save the cropped pixel group so we can convert it to imread object
        pixelGroup.save('pixelGroup.jpg')
        imgPixelGroup = imread('pixelGroup.jpg')

        r, g, b = [], [], [] # set up lists of r, g, b values
        for line in imgPixelGroup:
            for pixel in line:
                # get the rgb values of each pixel
                temp_r, temp_g, temp_b = pixel
                r.append(temp_r)
                g.append(temp_g)
                b.append(temp_b)

        df = pandas.DataFrame({
            'red': r,
            'green': g,
            'blue': b,
            })

        df_w = whiten(df)

        # get the cluster center (no need for distortion)
        cluster_center, distortion = kmeans(df_w, 1)


        color = []
        # standard deviation
        r_std, g_std, b_std = df[['red', 'green', 'blue']].std(ddof = 0)

        scaled_r, scaled_g, scaled_b = cluster_center[0]

        # adjust the rgb values scaling from 0 to 1
        color.append((
            scaled_r * std(r_std) / 255,
            scaled_g * std(g_std) / 255,
            scaled_b * std(b_std) / 255,
            ))

        cellColor = matplotlib.colors.to_hex(color[0]) # convert to hexadecimal
        # prepare the color of the cell
        cellFill = openpyxl.styles.PatternFill(
            fgColor = f'{cellColor[1:]}',
            fill_type = 'solid',
            )
        columnLetter = openpyxl.utils.get_column_letter(j + 1)
        rowNumber = i + 1
        # color the cell
        sheet[f'{columnLetter}' + f'{rowNumber}'].fill = cellFill


# delete pixelGroup.jpg
os.remove('pixelGroup.jpg')


# save the excel file
dotIndex = imageFileName.index('.')
excelFileName = sys.argv[1][:dotIndex]
wb.save(f'{excelFileName}.xlsx')